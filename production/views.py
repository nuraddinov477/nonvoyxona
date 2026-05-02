from django.db import models
from django.http import HttpResponse
from io import BytesIO
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import (RawMaterialCategory, RawMaterial, RawMaterialIncome,
                     Product, Recipe, DailyProduction, FinishedProductStock)
from .serializers import (RawMaterialCategorySerializer, RawMaterialSerializer,
                          RawMaterialIncomeSerializer, ProductSerializer,
                          RecipeSerializer, DailyProductionSerializer,
                          FinishedProductStockSerializer)


class RawMaterialCategoryViewSet(viewsets.ModelViewSet):
    queryset = RawMaterialCategory.objects.all()
    serializer_class = RawMaterialCategorySerializer


class RawMaterialViewSet(viewsets.ModelViewSet):
    queryset = RawMaterial.objects.select_related('category').all()
    serializer_class = RawMaterialSerializer
    filterset_fields = ['category']
    search_fields = ['name']

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Kam qolgan xom ashyolar"""
        materials = self.queryset.filter(quantity__lte=models.F('min_quantity'))
        return Response(self.get_serializer(materials, many=True).data)


class RawMaterialIncomeViewSet(viewsets.ModelViewSet):
    queryset = RawMaterialIncome.objects.select_related('material').all()
    serializer_class = RawMaterialIncomeSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.prefetch_related('recipes', 'stock').all()
    serializer_class = ProductSerializer
    filterset_fields = ['is_active']
    search_fields = ['name']


class RecipeViewSet(viewsets.ModelViewSet):
    queryset = Recipe.objects.select_related('product', 'material').all()
    serializer_class = RecipeSerializer
    filterset_fields = ['product']


class DailyProductionViewSet(viewsets.ModelViewSet):
    queryset = DailyProduction.objects.select_related('product', 'baker').all()
    serializer_class = DailyProductionSerializer
    filterset_fields = ['date', 'product', 'is_processed']

    def perform_create(self, serializer):
        production = serializer.save(baker=self.request.user)
        # Xom ashyolarni hisobdan chiqarish
        production.process_raw_materials()
        # Tayyor mahsulot omboriga qo'shish
        stock, _ = FinishedProductStock.objects.get_or_create(product=production.product)
        stock.quantity += production.quantity
        stock.save()


class FinishedProductStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FinishedProductStock.objects.select_related('product').all()
    serializer_class = FinishedProductStockSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_materials_excel(request):
    """Xom ashyo qoldiqlarini Excel'ga eksport"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    materials = RawMaterial.objects.select_related('category').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Xom ashyolar"

    headers = ['#', 'Nomi', 'Kategoriya', "O'lchov", 'Mavjud', 'Min. miqdor', "1 birlik narxi", "Jami qiymati", 'Holat']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='8B4513')
        cell.alignment = Alignment(horizontal='center')

    total_value = 0
    for i, m in enumerate(materials, 1):
        value = float(m.quantity) * float(m.price_per_unit)
        is_low = float(m.quantity) <= float(m.min_quantity)
        ws.append([
            i, m.name, m.category.name if m.category else '-', m.unit,
            float(m.quantity), float(m.min_quantity), float(m.price_per_unit),
            value, 'KAM!' if is_low else 'Yetarli'
        ])
        total_value += value

    ws.append([])
    ws.append(['', '', '', '', '', '', 'JAMI:', total_value, ''])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    widths = [5, 25, 15, 10, 12, 12, 14, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="xom_ashyolar.xlsx"'
    return response
